import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import PatternFill
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer
from sklearn.ensemble import RandomForestRegressor


def targeted_missforest_imputation_fixed(file_path):
    df_original = pd.read_excel(file_path)
    df_working = df_original.copy()  # 用于计算的工作副本

    # 用 openpyxl 记录颜色坐标
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print(f"读取文件失败: {e}")
        return

    # 确定受保护列
    protected_cols = ['样本ID', '高血脂症二分类标签', '血脂异常分型标签（确诊病例）']
    protected_indices = [df_original.columns.get_loc(col) for col in protected_cols if col in df_original.columns]

    # 定义样式
    fixed_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 修复成功：浅绿
    clear_fill = PatternFill(fill_type=None)  # 清除颜色：无色

    # 寻找所有带背景色的单元格
    outlier_coords = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        for c_idx, cell in enumerate(row):
            fill = cell.fill
            if fill and fill.start_color and fill.start_color.rgb:
                color_hex = str(fill.start_color.rgb)
                # 识别非白底填充
                if color_hex not in ['00000000', 'FFFFFFFF', 'System Window']:
                    outlier_coords.append((r_idx, c_idx))

    if not outlier_coords:
        print(" 未检测到任何异常颜色标记，无需处理。")
        return

    # 在工作副本中将异常生理指标标记为 NaN
    actual_impute_count = 0
    for r, c in outlier_coords:
        if c not in protected_indices:
            df_working.iat[r, c] = np.nan
            actual_impute_count += 1

    # 随机森林多重插补计算
    imputer = IterativeImputer(
        estimator=RandomForestRegressor(n_estimators=50, random_state=42),
        max_iter=10,
        random_state=42
    )
    # 计算完成后，df_imputed 包含了修正后的新值和原来的 ID/标签
    df_imputed = pd.DataFrame(imputer.fit_transform(df_working), columns=df_working.columns)

    # 精准写回值与格式还原
    for r, c in outlier_coords:
        excel_row = r + 2
        excel_col = c + 1

        # 获取列名以便判断类型
        col_name = df_original.columns[c]

        if c in protected_indices:
            original_val = df_original.iat[r, c]
            # 确保 ID 和 标签 保持为整数格式
            ws.cell(row=excel_row, column=excel_col).value = int(original_val)
            ws.cell(row=excel_row, column=excel_col).fill = clear_fill
        else:
            # 生理指标列，写入修正后的插补值，并标记为浅绿
            imputed_val = df_imputed.iat[r, c]
            # 保持生化指标的浮点精度
            ws.cell(row=excel_row, column=excel_col).value = imputed_val
            ws.cell(row=excel_row, column=excel_col).fill = fixed_fill

    # 导出结果
    dir_name = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    file_name, _ = os.path.splitext(base_name)
    export_path = os.path.join(dir_name, f"{file_name}_异常值处理.xlsx")

    try:
        wb.save(export_path)
        print(f"原始检测到异常坐标：{len(outlier_coords)} 处")
        print(f"随机森林多重插补")
        print(f"已保存至：\n{export_path}")
    except PermissionError:
        print(f"导出失败：请关闭已打开的 Excel 文件后重试。")

if __name__ == '__main__':
    # 路径
    TARGET_FILE_PATH = r"C:\Users\jack\Desktop\2026年第十六届MathorCup数学应用挑战赛赛题\C题\附件1：样例数据_多变量检查异常值.xlsx"
    targeted_missforest_imputation_fixed(TARGET_FILE_PATH)